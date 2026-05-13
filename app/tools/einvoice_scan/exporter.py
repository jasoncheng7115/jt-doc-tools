"""匯出 buffer 內發票資料 — CSV / XLSX / JSON.

設計：
- CSV：UTF-8 BOM (Excel 開繁中正確) + apply field_formats
- XLSX：標題列藍底白字粗體 freeze、欄寬 18 / 60 (備註)、apply field_formats
- JSON：永遠用內部標準格式（ISO 日期 / 整數金額 / compact 號碼），
        ignore field_formats — 給程式讀的契約

入口 build_export(invoices, columns, column_order, field_formats, fmt) 回 (bytes, mimetype, filename)
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Optional

from .formatters import apply_format
from .settings import FIELD_DEFINITIONS

_FIELD_DEF_BY_ID = {f["id"]: f for f in FIELD_DEFINITIONS}


def _resolve_columns(visible_columns: Optional[list[str]],
                     column_order: Optional[list[str]]) -> list[str]:
    """決定匯出哪些欄位 + 順序。預設全欄位 by FIELD_DEFINITIONS 順序。"""
    if not visible_columns:
        visible_columns = [f["id"] for f in FIELD_DEFINITIONS if f["default_visible"]]
    if not column_order:
        column_order = [f["id"] for f in FIELD_DEFINITIONS]
    return [c for c in column_order if c in visible_columns and c in _FIELD_DEF_BY_ID]


def _label(field_id: str, export_labels: Optional[dict] = None) -> str:
    """欄位顯示文字：使用者 export_labels 覆寫優先，否則用內建 label。"""
    if export_labels and isinstance(export_labels, dict):
        v = export_labels.get(field_id)
        if isinstance(v, str) and v.strip():
            return v.strip()
    d = _FIELD_DEF_BY_ID.get(field_id)
    return d["label"] if d else field_id


def _row_value(invoice: dict, field_id: str, row_index: int, field_formats: dict):
    """取一個欄位的顯示值（已 apply format）。"""
    if field_id == "seq":
        return row_index + 1
    if field_id == "tax":
        total = invoice.get("amount_total")
        untaxed = invoice.get("amount_untaxed")
        v = (total - untaxed) if (total is not None and untaxed is not None) else None
        return apply_format("tax", v, field_formats)
    if field_id == "items":
        items = invoice.get("items")
        if isinstance(items, list):
            return " / ".join(items)
        return ""
    return apply_format(field_id, invoice.get(field_id), field_formats)


def export_csv(invoices: list[dict], columns: list[str],
               field_formats: dict,
               export_labels: Optional[dict] = None) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow([_label(c, export_labels) for c in columns])
    for i, inv in enumerate(invoices):
        writer.writerow([_row_value(inv, c, i, field_formats) for c in columns])
    # BOM 給 Excel 開繁中正確
    return ("﻿" + buf.getvalue()).encode("utf-8")


def export_xlsx(invoices: list[dict], columns: list[str],
                field_formats: dict,
                export_labels: Optional[dict] = None) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl 未安裝，無法匯出 xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "einvoices"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    hdr_align = Alignment(horizontal="center", vertical="center")

    # Header
    for col_idx, c in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_label(c, export_labels))
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        # 欄寬：備註寬 60、品項寬 50、其他 18
        if c == "note":
            ws.column_dimensions[cell.column_letter].width = 60
        elif c == "items":
            ws.column_dimensions[cell.column_letter].width = 50
        else:
            ws.column_dimensions[cell.column_letter].width = 18

    # Rows
    for i, inv in enumerate(invoices):
        for col_idx, c in enumerate(columns, start=1):
            v = _row_value(inv, c, i, field_formats)
            # 序號用 int，其他都當字串（已 format）以保留千分位 / 民國格式
            if c == "seq":
                ws.cell(row=i + 2, column=col_idx, value=int(v))
            else:
                ws.cell(row=i + 2, column=col_idx, value=v if v != "" else None)

    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def export_json(invoices: list[dict], columns: list[str],
                field_formats: dict) -> bytes:
    """JSON 永遠用內部標準格式 — 忽略 field_formats。

    輸出每筆 invoice 是 dict（key = field_id），值為 raw 內部格式：
    - invoice_number: compact 字串
    - date: ISO 'YYYY-MM-DD'
    - amount_*: int
    - scanned_at: ISO 8601
    - items: list 或 缺
    """
    out = []
    for i, inv in enumerate(invoices):
        entry = {}
        for c in columns:
            if c == "seq":
                entry[c] = i + 1
            elif c == "tax":
                total = inv.get("amount_total")
                untaxed = inv.get("amount_untaxed")
                entry[c] = (total - untaxed) if (total is not None and untaxed is not None) else None
            else:
                entry[c] = inv.get(c)
        out.append(entry)
    return json.dumps({"invoices": out, "exported_at": datetime.now().isoformat()},
                      ensure_ascii=False, indent=2).encode("utf-8")


def build_export(invoices: list[dict], visible_columns: list[str],
                 column_order: list[str], field_formats: dict,
                 fmt: str,
                 export_labels: Optional[dict] = None) -> tuple[bytes, str, str]:
    """統一入口。回 (data, mimetype, filename_suffix).

    export_labels: dict[field_id, str] 自訂欄位標題（供匯入其他系統用）。
    """
    columns = _resolve_columns(visible_columns, column_order)
    fmt = (fmt or "").lower()
    if fmt == "csv":
        return (export_csv(invoices, columns, field_formats, export_labels),
                "text/csv; charset=utf-8", "csv")
    if fmt == "xlsx":
        return (export_xlsx(invoices, columns, field_formats, export_labels),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "xlsx")
    if fmt == "json":
        return (export_json(invoices, columns, field_formats),
                "application/json; charset=utf-8", "json")
    if fmt == "ods":
        return (export_ods(invoices, columns, field_formats, export_labels),
                "application/vnd.oasis.opendocument.spreadsheet", "ods")
    if fmt == "xml":
        return (export_xml(invoices, columns, field_formats),
                "application/xml; charset=utf-8", "xml")
    if fmt == "txt":
        return (export_txt(invoices, columns, field_formats, export_labels),
                "text/plain; charset=utf-8", "txt")
    if fmt == "md":
        return (export_md(invoices, columns, field_formats, export_labels),
                "text/markdown; charset=utf-8", "md")
    raise ValueError(f"不支援的匯出格式：{fmt}")


def export_ods(invoices: list[dict], columns: list[str],
               field_formats: dict,
               export_labels: Optional[dict] = None) -> bytes:
    """ODS — LibreOffice Calc 原生格式。用 odfpy。"""
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.style import (Style, TextProperties, TableColumnProperties,
                               TableCellProperties)
        from odf.table import Table, TableColumn, TableRow, TableCell
        from odf.text import P
    except ImportError:
        raise RuntimeError("odfpy 未安裝，無法匯出 ods")
    doc = OpenDocumentSpreadsheet()
    # Header style: 藍底白字粗體
    hdr_style = Style(name="HdrCell", family="table-cell")
    hdr_style.addElement(TextProperties(fontweight="bold", color="#FFFFFF"))
    hdr_style.addElement(TableCellProperties(backgroundcolor="#2563EB"))
    doc.styles.addElement(hdr_style)
    table = Table(name="einvoices")
    for _ in columns:
        table.addElement(TableColumn())
    # Header row
    hdr = TableRow()
    for c in columns:
        cell = TableCell(stylename=hdr_style)
        cell.addElement(P(text=_label(c, export_labels)))
        hdr.addElement(cell)
    table.addElement(hdr)
    # Data rows
    for i, inv in enumerate(invoices):
        tr = TableRow()
        for c in columns:
            v = _row_value(inv, c, i, field_formats)
            cell = TableCell()
            cell.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(cell)
        table.addElement(tr)
    doc.spreadsheet.addElement(table)
    buf = io.BytesIO()
    doc.write(buf)
    return buf.getvalue()


def _xml_escape(s):
    """最小 XML escape — 處理 & < > " '"""
    if s is None:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&apos;"))


def export_xml(invoices: list[dict], columns: list[str],
               field_formats: dict) -> bytes:
    """XML — 簡單結構：<invoices><invoice><field>value</field>...</invoice>...</invoices>"""
    lines = ['<?xml version="1.0" encoding="utf-8"?>', "<invoices>"]
    for i, inv in enumerate(invoices):
        lines.append("  <invoice>")
        for c in columns:
            v = _row_value(inv, c, i, field_formats)
            lines.append(f"    <{c}>{_xml_escape(v)}</{c}>")
        lines.append("  </invoice>")
    lines.append("</invoices>")
    return ("\n".join(lines) + "\n").encode("utf-8")


def export_txt(invoices: list[dict], columns: list[str],
               field_formats: dict,
               export_labels: Optional[dict] = None) -> bytes:
    """純文字 — 直式對齊欄位（適合貼到 email / log）。"""
    # 計算每欄寬度
    headers = [_label(c, export_labels) for c in columns]
    widths = [_display_width(h) for h in headers]
    rows = []
    for i, inv in enumerate(invoices):
        row = [str(_row_value(inv, c, i, field_formats)) for c in columns]
        rows.append(row)
        for j, v in enumerate(row):
            widths[j] = max(widths[j], _display_width(v))
    out = []
    out.append("  ".join(_pad(headers[j], widths[j]) for j in range(len(columns))))
    out.append("  ".join("-" * widths[j] for j in range(len(columns))))
    for row in rows:
        out.append("  ".join(_pad(row[j], widths[j]) for j in range(len(columns))))
    return ("\n".join(out) + "\n").encode("utf-8")


def export_md(invoices: list[dict], columns: list[str],
              field_formats: dict,
              export_labels: Optional[dict] = None) -> bytes:
    """Markdown — GFM table 格式。"""
    headers = [_label(c, export_labels) for c in columns]
    out = ["| " + " | ".join(_md_escape(h) for h in headers) + " |"]
    out.append("| " + " | ".join("---" for _ in columns) + " |")
    for i, inv in enumerate(invoices):
        row = [_md_escape(_row_value(inv, c, i, field_formats)) for c in columns]
        out.append("| " + " | ".join(row) + " |")
    return ("\n".join(out) + "\n").encode("utf-8")


def _md_escape(s):
    if s is None:
        return ""
    # Markdown table 內 | 需 escape；換行替換成空白
    return str(s).replace("|", "\\|").replace("\n", " ").replace("\r", "")


def _display_width(s):
    """大略估字寬：CJK = 2，其他 = 1。"""
    if s is None:
        return 0
    n = 0
    for ch in str(s):
        cp = ord(ch)
        # 簡單判定 CJK / 全形
        if (0x3000 <= cp <= 0x9FFF) or (0xFF00 <= cp <= 0xFFEF):
            n += 2
        else:
            n += 1
    return n


def _pad(s, width):
    s = str(s) if s is not None else ""
    actual = _display_width(s)
    return s + " " * max(0, width - actual)
