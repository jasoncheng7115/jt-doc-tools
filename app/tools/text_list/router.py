"""清單處理 endpoints — 把多行文字當清單做排序 / 去重 / 篩選 / 取頭尾。

設計為「pipeline of ops」：前端傳一個 `ops` 陣列，後端依序套用，
每個 op 的輸出餵給下一個 op。新增功能只要加一個 `_op_*` 函式 +
在 `_OPS` dict 註冊，前端 UI 加勾選。

支援的輸入格式：
- 貼上純文字（textarea）
- 上傳 .txt / .csv / .tsv / .md / .log / .json：UTF-8 讀檔 splitlines
- 上傳 .xlsx / .ods：openpyxl / odfpy 取第一個工作表的所有非空 cell
  （扁平化為一維清單，row-major）
- 上傳 .docx / .odt：python-docx / odfpy 取段落，每段一行
- 上傳 .pdf：PyMuPDF get_text 後 splitlines

輸出：
- JSON：{lines: [...], count, dropped, ops_applied}
- 下載 .txt / .csv / .xlsx
"""
from __future__ import annotations

import csv
import io
import random
import re
from typing import Any, Callable

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse

from ...core.http_utils import content_disposition

router = APIRouter()

_MAX_INPUT_BYTES = 8 * 1024 * 1024   # 8 MiB cap per request body / file
_MAX_LINES = 500_000                  # 上限避免 DoS（500k 行已是 ~50MB 純文字）
_MAX_FILTER_REGEX_LEN = 1000          # regex 上限避免 ReDoS pattern


# ──────────────────────────────────────────────────────────────────────
# 檔案萃取 — 把上傳檔案轉成 list[str]
# ──────────────────────────────────────────────────────────────────────

def _extract_text_lines(filename: str, data: bytes) -> list[str]:
    """依副檔名選擇萃取方式；失敗回 HTTPException(400)。

    一行一筆原則：
    - 純文字 / CSV / TSV / log / json：以 splitlines() 拆行
    - xlsx / ods：每個非空 cell 一行（row-major，扁平化）
    - docx / odt：每個非空段落一行
    - pdf：PyMuPDF get_text 後 splitlines
    """
    name = (filename or "").lower()
    suffix = name.rsplit(".", 1)[-1] if "." in name else ""

    text_suffixes = {"txt", "csv", "tsv", "md", "markdown", "log", "json", "yaml", "yml", "ini", "conf"}
    if suffix in text_suffixes or not suffix:
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            try:
                text = data.decode("big5")  # 台灣常見舊檔案
            except UnicodeDecodeError:
                raise HTTPException(400, f"檔案不是 UTF-8 或 Big5 編碼，無法讀取")
        return text.splitlines()

    if suffix in ("xlsx", "ods"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(500, "openpyxl 未安裝，無法讀 xlsx/ods")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(400, f"無法解析試算表：{e}")
        ws = wb.active
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell).strip()
                if s:
                    lines.append(s)
                    if len(lines) >= _MAX_LINES:
                        return lines
        return lines

    if suffix == "docx":
        try:
            import docx
        except ImportError:
            raise HTTPException(500, "python-docx 未安裝，無法讀 docx")
        try:
            d = docx.Document(io.BytesIO(data))
        except Exception as e:
            raise HTTPException(400, f"無法解析 docx：{e}")
        return [p.text for p in d.paragraphs if p.text and p.text.strip()]

    if suffix == "odt":
        try:
            from odf import opendocument, text as odf_text, teletype
        except ImportError:
            raise HTTPException(500, "odfpy 未安裝，無法讀 odt")
        try:
            doc = opendocument.load(io.BytesIO(data))
        except Exception as e:
            raise HTTPException(400, f"無法解析 odt：{e}")
        lines: list[str] = []
        for p in doc.getElementsByType(odf_text.P):
            s = teletype.extractText(p).strip()
            if s:
                lines.append(s)
        return lines

    if suffix == "pdf":
        try:
            import fitz
        except ImportError:
            raise HTTPException(500, "PyMuPDF 未安裝，無法讀 pdf")
        try:
            d = fitz.open(stream=data, filetype="pdf")
        except Exception as e:
            raise HTTPException(400, f"無法解析 pdf：{e}")
        lines: list[str] = []
        for page in d:
            for ln in page.get_text().splitlines():
                if ln.strip():
                    lines.append(ln)
                    if len(lines) >= _MAX_LINES:
                        d.close()
                        return lines
        d.close()
        return lines

    raise HTTPException(400, f"不支援的副檔名 .{suffix}（支援：.txt .csv .tsv .md .log .json .xlsx .ods .docx .odt .pdf）")


# ──────────────────────────────────────────────────────────────────────
# 操作 pipeline — 每個 op 取 (lines, params) 回 lines
# ──────────────────────────────────────────────────────────────────────

def _natural_key(s: str) -> tuple:
    """把 'item10' 拆成 ['item', 10] 讓 'item2' < 'item10'。"""
    parts = re.split(r"(\d+)", s or "")
    return tuple((int(p) if p.isdigit() else p.lower()) for p in parts if p)


def _op_trim(lines: list[str], _params: dict) -> list[str]:
    return [ln.strip() for ln in lines]


def _op_drop_empty(lines: list[str], _params: dict) -> list[str]:
    return [ln for ln in lines if ln.strip()]


def _op_dedup(lines: list[str], params: dict) -> list[str]:
    keep = params.get("keep", "first")
    case_insensitive = bool(params.get("case_insensitive", False))
    key_fn = (lambda s: s.lower()) if case_insensitive else (lambda s: s)
    if keep == "last":
        seen: dict[str, int] = {}
        for i, ln in enumerate(lines):
            seen[key_fn(ln)] = i
        keep_idx = set(seen.values())
        return [ln for i, ln in enumerate(lines) if i in keep_idx]
    if keep == "count":
        # 「列出每筆 + 出現次數」，依首次出現順序
        from collections import Counter, OrderedDict
        order = OrderedDict()
        for ln in lines:
            k = key_fn(ln)
            if k not in order:
                order[k] = ln  # 保留首次原文
        counts = Counter(key_fn(ln) for ln in lines)
        return [f"{counts[k]}\t{order[k]}" for k in order]
    # default: keep first
    seen_set: set[str] = set()
    out: list[str] = []
    for ln in lines:
        k = key_fn(ln)
        if k in seen_set:
            continue
        seen_set.add(k)
        out.append(ln)
    return out


def _op_sort(lines: list[str], params: dict) -> list[str]:
    order = params.get("order", "asc")
    natural = bool(params.get("natural", False))
    case_insensitive = bool(params.get("case_insensitive", False))
    if natural:
        key_fn = _natural_key
    elif case_insensitive:
        key_fn = lambda s: s.lower()  # noqa: E731
    else:
        key_fn = lambda s: s  # noqa: E731
    return sorted(lines, key=key_fn, reverse=(order == "desc"))


def _op_reverse(lines: list[str], _params: dict) -> list[str]:
    return list(reversed(lines))


def _op_shuffle(lines: list[str], params: dict) -> list[str]:
    seed = params.get("seed")
    rng = random.Random(seed) if seed is not None else random.Random()
    out = list(lines)
    rng.shuffle(out)
    return out


def _op_case(lines: list[str], params: dict) -> list[str]:
    mode = params.get("mode", "lower")
    if mode == "lower":
        return [ln.lower() for ln in lines]
    if mode == "upper":
        return [ln.upper() for ln in lines]
    if mode == "title":
        return [ln.title() for ln in lines]
    return lines


def _op_filter(lines: list[str], params: dict) -> list[str]:
    mode = params.get("mode", "include")  # include / exclude
    pattern = (params.get("pattern") or "")
    is_regex = bool(params.get("regex", False))
    case_insensitive = bool(params.get("case_insensitive", False))
    if not pattern:
        return lines
    if is_regex:
        if len(pattern) > _MAX_FILTER_REGEX_LEN:
            raise HTTPException(400, f"regex 過長（>{_MAX_FILTER_REGEX_LEN}）")
        flags = re.IGNORECASE if case_insensitive else 0
        try:
            rx = re.compile(pattern, flags)
        except re.error as e:
            raise HTTPException(400, f"regex 編譯失敗：{e}")
        match_fn = lambda s: rx.search(s) is not None  # noqa: E731
    else:
        if case_insensitive:
            needle = pattern.lower()
            match_fn = lambda s: needle in s.lower()  # noqa: E731
        else:
            match_fn = lambda s: pattern in s  # noqa: E731
    if mode == "exclude":
        return [ln for ln in lines if not match_fn(ln)]
    return [ln for ln in lines if match_fn(ln)]


def _op_head(lines: list[str], params: dict) -> list[str]:
    try:
        n = int(params.get("n", 10))
    except (TypeError, ValueError):
        raise HTTPException(400, "head: n 必須是整數")
    if n < 0:
        raise HTTPException(400, "head: n 不可為負")
    return lines[:n]


def _op_tail(lines: list[str], params: dict) -> list[str]:
    try:
        n = int(params.get("n", 10))
    except (TypeError, ValueError):
        raise HTTPException(400, "tail: n 必須是整數")
    if n < 0:
        raise HTTPException(400, "tail: n 不可為負")
    return lines[-n:] if n > 0 else []


def _op_prefix_suffix(lines: list[str], params: dict) -> list[str]:
    prefix = params.get("prefix", "") or ""
    suffix = params.get("suffix", "") or ""
    if not prefix and not suffix:
        return lines
    return [f"{prefix}{ln}{suffix}" for ln in lines]


_OPS: dict[str, Callable[[list[str], dict], list[str]]] = {
    "trim": _op_trim,
    "drop_empty": _op_drop_empty,
    "dedup": _op_dedup,
    "sort": _op_sort,
    "reverse": _op_reverse,
    "shuffle": _op_shuffle,
    "case": _op_case,
    "filter": _op_filter,
    "head": _op_head,
    "tail": _op_tail,
    "wrap": _op_prefix_suffix,
}


def _apply_pipeline(lines: list[str], ops: list[dict]) -> tuple[list[str], list[str]]:
    """套用 ops 到 lines，回 (final_lines, applied_op_names)。"""
    applied = []
    for op_spec in ops or []:
        if not isinstance(op_spec, dict):
            continue
        op_name = op_spec.get("op")
        if not op_name or op_name not in _OPS:
            continue
        params = {k: v for k, v in op_spec.items() if k != "op"}
        lines = _OPS[op_name](lines, params)
        applied.append(op_name)
    return lines, applied


# ──────────────────────────────────────────────────────────────────────
# Endpoints
# ──────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def index(request: Request):
    templates = request.app.state.templates
    return templates.TemplateResponse("text_list.html", {"request": request})


@router.post("/process")
async def process(request: Request):
    """JSON in/out — 主要的後端處理 endpoint。"""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "invalid JSON body")
    text = body.get("text", "")
    if not isinstance(text, str):
        raise HTTPException(400, "text 必須是字串")
    if len(text.encode("utf-8")) > _MAX_INPUT_BYTES:
        raise HTTPException(413, f"輸入超過 {_MAX_INPUT_BYTES // 1024 // 1024} MiB")
    ops = body.get("ops", [])
    if not isinstance(ops, list):
        raise HTTPException(400, "ops 必須是陣列")
    lines = text.splitlines()
    if len(lines) > _MAX_LINES:
        raise HTTPException(413, f"超過 {_MAX_LINES:,} 行上限")
    original_count = len(lines)
    out, applied = _apply_pipeline(lines, ops)
    return JSONResponse({
        "lines": out,
        "count": len(out),
        "original_count": original_count,
        "dropped": original_count - len(out),
        "ops_applied": applied,
    })


@router.post("/upload")
async def upload(request: Request, file: UploadFile = File(...)):
    """上傳檔案 → 萃取為 lines → 直接套 pipeline（從 form 讀 ops JSON）。

    回 JSON 同 /process。前端可選擇先上傳再套 ops，或直接傳 ops 一次處理完。
    """
    form = await request.form()
    ops_raw = form.get("ops") or "[]"
    try:
        import json
        ops = json.loads(ops_raw)
    except Exception:
        raise HTTPException(400, "ops form field 必須是 JSON 陣列")
    if not isinstance(ops, list):
        raise HTTPException(400, "ops 必須是陣列")

    data = await file.read()
    if len(data) > _MAX_INPUT_BYTES:
        raise HTTPException(413, f"檔案超過 {_MAX_INPUT_BYTES // 1024 // 1024} MiB 上限")
    lines = _extract_text_lines(file.filename or "", data)
    if len(lines) > _MAX_LINES:
        raise HTTPException(413, f"超過 {_MAX_LINES:,} 行上限")
    original_count = len(lines)
    out, applied = _apply_pipeline(lines, ops)
    return JSONResponse({
        "lines": out,
        "count": len(out),
        "original_count": original_count,
        "dropped": original_count - len(out),
        "ops_applied": applied,
        "filename": file.filename,
    })


@router.post("/export/{fmt}")
async def export(fmt: str, request: Request):
    """把處理結果下載成檔案。fmt: txt / csv / xlsx。"""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "invalid JSON body")
    lines = body.get("lines", [])
    if not isinstance(lines, list):
        raise HTTPException(400, "lines 必須是陣列")
    if len(lines) > _MAX_LINES:
        raise HTTPException(413, f"超過 {_MAX_LINES:,} 行上限")
    # 一律強轉字串避免 export 階段炸
    lines = [str(ln) if ln is not None else "" for ln in lines]

    fmt = (fmt or "").lower()
    base = body.get("filename") or "list"
    # 移除副檔名
    if "." in base:
        base = base.rsplit(".", 1)[0]
    base = re.sub(r"[\\/:*?\"<>|]", "_", base)[:80] or "list"

    if fmt == "txt":
        data = ("\n".join(lines) + "\n").encode("utf-8")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": content_disposition(f"{base}.txt")},
        )

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        for ln in lines:
            # 若是 dedup-count 結果（"count\toriginal"），拆兩欄；否則單欄
            if "\t" in ln:
                parts = ln.split("\t", 1)
                writer.writerow(parts)
            else:
                writer.writerow([ln])
        data = ("﻿" + buf.getvalue()).encode("utf-8")  # BOM 給 Excel 開繁中正確
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": content_disposition(f"{base}.csv")},
        )

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(500, "openpyxl 未安裝，無法匯出 xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "list"
        # 偵測是否為 dedup-count 結果（每行都含 \t）
        is_count = lines and all("\t" in ln for ln in lines)
        if is_count:
            ws.cell(row=1, column=1, value="次數").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=2, value="內容").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="2563EB")
            ws.cell(row=1, column=2).fill = PatternFill("solid", fgColor="2563EB")
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")
            for i, ln in enumerate(lines, start=2):
                count_str, content = ln.split("\t", 1)
                try:
                    ws.cell(row=i, column=1, value=int(count_str))
                except ValueError:
                    ws.cell(row=i, column=1, value=count_str)
                ws.cell(row=i, column=2, value=content)
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 60
            ws.freeze_panes = "A2"
        else:
            ws.cell(row=1, column=1, value="內容").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="2563EB")
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            for i, ln in enumerate(lines, start=2):
                ws.cell(row=i, column=1, value=ln)
            ws.column_dimensions["A"].width = 60
            ws.freeze_panes = "A2"
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": content_disposition(f"{base}.xlsx")},
        )

    raise HTTPException(400, f"不支援的格式：{fmt}（支援：txt / csv / xlsx）")


@router.post("/api/text-list")
async def api_text_list(request: Request):
    """Public alias — 同 /process。給 REST API caller 用。"""
    return await process(request)
