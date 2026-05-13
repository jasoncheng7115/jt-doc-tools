"""Tests for text-list tool — pipeline ops, file extraction, export formats.

Mirrors the test pattern of tests/test_text_diff.py: hits endpoints via the
TestClient using the real FastAPI app, asserts JSON shape & file content.
"""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.tools.text_list.router import (
    _apply_pipeline,
    _natural_key,
    _OPS,
)


client = TestClient(app)


# ─── Unit: ops pipeline ──────────────────────────────────────────────

def test_op_registry_has_expected():
    assert "sort" in _OPS
    assert "dedup" in _OPS
    assert "filter" in _OPS
    assert len(_OPS) >= 10  # 至少 11 個 ops


def test_pipeline_sort_dedup():
    lines = ["banana", "apple", "cherry", "apple"]
    out, applied = _apply_pipeline(lines, [
        {"op": "dedup"},
        {"op": "sort", "order": "asc"},
    ])
    assert out == ["apple", "banana", "cherry"]
    assert applied == ["dedup", "sort"]


def test_pipeline_dedup_keep_last():
    lines = ["a", "b", "a", "c", "b"]
    out, _ = _apply_pipeline(lines, [{"op": "dedup", "keep": "last"}])
    assert out == ["a", "c", "b"]


def test_pipeline_dedup_count():
    lines = ["a", "b", "a", "c", "b", "a"]
    out, _ = _apply_pipeline(lines, [{"op": "dedup", "keep": "count"}])
    # 順序依首次出現：a (3), b (2), c (1)
    assert out == ["3\ta", "2\tb", "1\tc"]


def test_pipeline_dedup_case_insensitive():
    lines = ["Apple", "apple", "APPLE", "banana"]
    out, _ = _apply_pipeline(lines, [{"op": "dedup", "case_insensitive": True}])
    assert out == ["Apple", "banana"]  # 第一個出現的保留


def test_natural_sort():
    lines = ["item10", "item2", "item1", "item20"]
    out, _ = _apply_pipeline(lines, [{"op": "sort", "natural": True}])
    assert out == ["item1", "item2", "item10", "item20"]


def test_pipeline_filter_include():
    lines = ["apple pie", "banana bread", "apple sauce"]
    out, _ = _apply_pipeline(lines, [
        {"op": "filter", "mode": "include", "pattern": "apple"},
    ])
    assert out == ["apple pie", "apple sauce"]


def test_pipeline_filter_regex():
    lines = ["abc", "a1b", "xyz", "a23b"]
    out, _ = _apply_pipeline(lines, [
        {"op": "filter", "mode": "include", "pattern": r"^a\d+b$", "regex": True},
    ])
    assert out == ["a1b", "a23b"]


def test_pipeline_filter_invalid_regex():
    """regex compile error → HTTP 400 (raised by op)."""
    from fastapi import HTTPException
    with pytest.raises(HTTPException) as ei:
        _apply_pipeline(["x"], [
            {"op": "filter", "mode": "include", "pattern": "(unclosed", "regex": True},
        ])
    assert ei.value.status_code == 400


def test_pipeline_head_tail():
    lines = list("abcdefghij")
    out, _ = _apply_pipeline(lines, [{"op": "head", "n": 3}])
    assert out == ["a", "b", "c"]
    out, _ = _apply_pipeline(lines, [{"op": "tail", "n": 2}])
    assert out == ["i", "j"]


def test_pipeline_reverse():
    out, _ = _apply_pipeline(["a", "b", "c"], [{"op": "reverse"}])
    assert out == ["c", "b", "a"]


def test_pipeline_case():
    out, _ = _apply_pipeline(["Hello", "WORLD"], [{"op": "case", "mode": "lower"}])
    assert out == ["hello", "world"]
    out, _ = _apply_pipeline(["Hello", "world"], [{"op": "case", "mode": "upper"}])
    assert out == ["HELLO", "WORLD"]


def test_pipeline_trim_drop_empty():
    out, _ = _apply_pipeline(
        ["  a  ", "", "   ", " b "],
        [{"op": "trim"}, {"op": "drop_empty"}],
    )
    assert out == ["a", "b"]


def test_pipeline_wrap_prefix_suffix():
    out, _ = _apply_pipeline(
        ["one", "two"],
        [{"op": "wrap", "prefix": "<", "suffix": ">"}],
    )
    assert out == ["<one>", "<two>"]


def test_pipeline_unknown_op_skipped():
    out, applied = _apply_pipeline(
        ["a", "b"],
        [{"op": "nonexistent"}, {"op": "reverse"}],
    )
    assert out == ["b", "a"]
    assert applied == ["reverse"]  # 未知 op skip 但不爆炸


def test_natural_key():
    # 保證 'item2' < 'item10'
    assert _natural_key("item2") < _natural_key("item10")
    assert _natural_key("a1b") < _natural_key("a2b")


# ─── HTTP: /process JSON ──────────────────────────────────────────────

def test_process_basic():
    r = client.post("/tools/text-list/process", json={
        "text": "banana\napple\ncherry\napple\n",
        "ops": [{"op": "dedup"}, {"op": "sort", "order": "asc"}],
    })
    assert r.status_code == 200
    j = r.json()
    assert j["lines"] == ["apple", "banana", "cherry"]
    assert j["count"] == 3
    assert j["original_count"] == 4
    assert j["ops_applied"] == ["dedup", "sort"]


def test_process_empty_input():
    r = client.post("/tools/text-list/process", json={"text": "", "ops": []})
    assert r.status_code == 200
    assert r.json()["count"] == 0


def test_process_invalid_text_type():
    r = client.post("/tools/text-list/process", json={"text": 123, "ops": []})
    assert r.status_code == 400


def test_process_invalid_ops_type():
    r = client.post("/tools/text-list/process", json={"text": "a", "ops": "not-a-list"})
    assert r.status_code == 400


def test_process_oversize_input():
    big = "x\n" * 600_000  # 超過 _MAX_LINES = 500k
    r = client.post("/tools/text-list/process", json={"text": big, "ops": []})
    assert r.status_code == 413


# ─── HTTP: /api/text-list public alias ────────────────────────────────

def test_api_alias_works():
    r = client.post("/tools/text-list/api/text-list", json={
        "text": "c\na\nb\n",
        "ops": [{"op": "sort"}],
    })
    assert r.status_code == 200
    assert r.json()["lines"] == ["a", "b", "c"]


# ─── HTTP: /upload file extraction ───────────────────────────────────

def test_upload_txt():
    r = client.post(
        "/tools/text-list/upload",
        files={"file": ("test.txt", b"line1\nline2\nline3\n", "text/plain")},
        data={"ops": "[]"},
    )
    assert r.status_code == 200
    j = r.json()
    assert j["lines"] == ["line1", "line2", "line3"]
    assert j["filename"] == "test.txt"


def test_upload_csv_with_ops():
    r = client.post(
        "/tools/text-list/upload",
        files={"file": ("data.csv", b"banana\napple\nbanana\n", "text/csv")},
        data={"ops": '[{"op":"dedup"},{"op":"sort"}]'},
    )
    assert r.status_code == 200
    assert r.json()["lines"] == ["apple", "banana"]


def test_upload_xlsx():
    """xlsx with two cells in two rows → two output lines (row-major)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "first"
    ws["A2"] = "second"
    ws["A3"] = "third"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        "/tools/text-list/upload",
        files={"file": ("data.xlsx", buf.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"ops": "[]"},
    )
    assert r.status_code == 200
    assert r.json()["lines"] == ["first", "second", "third"]


def test_upload_unsupported_format():
    r = client.post(
        "/tools/text-list/upload",
        files={"file": ("test.exe", b"\x4d\x5a\x90\x00", "application/octet-stream")},
        data={"ops": "[]"},
    )
    assert r.status_code == 400
    assert "不支援" in r.json()["detail"] or "supported" in r.json()["detail"].lower()


def test_upload_invalid_ops_json():
    r = client.post(
        "/tools/text-list/upload",
        files={"file": ("a.txt", b"hi\n", "text/plain")},
        data={"ops": "not-json"},
    )
    assert r.status_code == 400


# ─── HTTP: /export download formats ──────────────────────────────────

def test_export_txt():
    r = client.post("/tools/text-list/export/txt", json={
        "lines": ["one", "two", "three"],
        "filename": "myfile.csv",  # 副檔名會被脫除
    })
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/plain")
    assert b"one\ntwo\nthree" in r.content
    cd = r.headers.get("content-disposition", "")
    assert "myfile.txt" in cd


def test_export_csv_includes_bom():
    """CSV 給 Excel 開繁中需要 BOM。"""
    r = client.post("/tools/text-list/export/csv", json={
        "lines": ["中文一", "中文二"],
    })
    assert r.status_code == 200
    # UTF-8 BOM = EF BB BF
    assert r.content.startswith(b"\xef\xbb\xbf")
    assert "中文一".encode("utf-8") in r.content


def test_export_csv_with_count_split():
    """dedup-count 的 'N\\tcontent' → CSV 拆兩欄。"""
    r = client.post("/tools/text-list/export/csv", json={
        "lines": ["3\tapple", "1\tbanana"],
    })
    assert r.status_code == 200
    text = r.content.decode("utf-8-sig")
    assert "3,apple" in text
    assert "1,banana" in text


def test_export_xlsx():
    r = client.post("/tools/text-list/export/xlsx", json={
        "lines": ["alpha", "beta", "gamma"],
        "filename": "list",
    })
    assert r.status_code == 200
    # PK = zip header (xlsx 是 zip)
    assert r.content.startswith(b"PK")
    # 驗證內容真的是有效的 xlsx
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws["A1"].value == "內容"  # header
    assert ws["A2"].value == "alpha"
    assert ws["A4"].value == "gamma"


def test_export_xlsx_count_format():
    """dedup-count 的 'N\\tcontent' → xlsx 兩欄含 header「次數 / 內容」。"""
    r = client.post("/tools/text-list/export/xlsx", json={
        "lines": ["5\tapple", "2\tbanana"],
    })
    assert r.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws["A1"].value == "次數"
    assert ws["B1"].value == "內容"
    assert ws["A2"].value == 5
    assert ws["B2"].value == "apple"


def test_export_unknown_format():
    r = client.post("/tools/text-list/export/exe", json={"lines": ["a"]})
    assert r.status_code == 400


def test_export_filename_sanitization():
    """檔名含 path separator 應被替換掉，避免 path traversal。"""
    r = client.post("/tools/text-list/export/txt", json={
        "lines": ["x"],
        "filename": "../../../etc/passwd",
    })
    assert r.status_code == 200
    cd = r.headers.get("content-disposition", "")
    # 不能含 ../ 或 / — sanitize 把 / 換成 _，避免 path traversal
    assert "../" not in cd
    assert "/etc/" not in cd
    assert "/passwd" not in cd


# ─── End-to-end: index page renders ──────────────────────────────────

def test_index_renders():
    r = client.get("/tools/text-list/")
    assert r.status_code == 200
    assert "清單處理" in r.text
    assert "全部複製" in r.text
